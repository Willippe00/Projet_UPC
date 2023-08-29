from git.Projet_UPC.Database.DataBaseManager import DataBaseManager
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

class presentationResultas:
    nom_fichier = "./sortie.xlsx"

    def __init__(self):
        self.DataBase = DataBaseManager()

        if os.path.exists(self.nom_fichier):
            os.remove(self.nom_fichier)
            print(f"Le fichier '{self.nom_fichier}' a été supprimé.")
        else:
            print(f"Le fichier '{self.nom_fichier}' n'existe pas.")


    def présenterCorespondance(self, RB_procuct):

        if os.path.exists(self.nom_fichier):
            print(f"Le fichier '{self.nom_fichier}' existe.")
        else:
            print(f"Le fichier '{self.nom_fichier}' n'existe pas. Création du fichier...")
            wb = Workbook()
            wb.save(self.nom_fichier)
            print(f"Le fichier '{self.nom_fichier}' a été créé.")

        wb = load_workbook(self.nom_fichier)
        ws = wb.active

        cellule = ws['A1']  # Départ à la cellule A1
        while cellule.value is not None:
            print(cellule.value)
            cellule = ws.cell(row=cellule.row + 1, column=1)

        cellule_origine = cellule

        cellule.value = RB_procuct
        ws.cell(row=cellule_origine.row + 1, column=1).value = "UPC"
        ws.cell(row=cellule_origine.row + 2, column=1).value = "image"
        ws.cell(row=cellule_origine.row + 3, column=1).value = "hyperlien"

        résultas_trier = self.récupérerCorespondance(RB_procuct)

        cellule_libre = self.trouver_premiere_cellule_vide_dans_rangée(ws, 1)

        for résulta in résultas_trier:
            numéro_ranger = cellule_origine.row + 1

            cellule_libre = self.trouver_premiere_cellule_vide_dans_rangée(ws, numéro_ranger)

            print("code upc:")
            print(résulta[1])
            if cellule_libre is not None:
                cellule_libre.value = résulta[1]  # Écrire le code UPC
                cellule_libre.fill = self.get_couleur_pourcentage(résulta[6])

                # Coordonnées de la cellule libre
                x = cellule_libre.column
                y = cellule_libre.row

                cellule_image = ws.cell(row=y + 1, column=x)
                try:
                  image = Image("./image/barcode/"+RB_procuct+"/"+résulta[1]+"/"+résulta[1] )
                  image.width = 100
                  image.height = 100

                  # Insérer l'image en spécifiant les coordonnées de la cellule libre
                  ws.add_image(image, f'{cellule_image.column_letter}{cellule_image.row}')

                  # Ajuster la hauteur de la rangée pour qu'elle corresponde à la hauteur de l'image
                  ws.row_dimensions[cellule_image.row].height = image.height

                  # Ajuster la largeur de la colonne pour qu'elle corresponde à la largeur de l'image
                  lettre_colonne = get_column_letter(cellule_image.column)
                  ws.column_dimensions[lettre_colonne].width = image.width / 7.5  # Vous pouvez ajuster le coefficient en fonction de vos besoins
                except Exception as e:
                  print("pas de d'image ")

                cellule_hyperlien = ws.cell(row=y + 2, column=x)
                cellule_hyperlien.value = "https://www.barcodelookup.com/"+résulta[1]
                cellule_hyperlien.hyperlink = "https://www.barcodelookup.com/"+résulta[1]
                cellule_hyperlien.style = "Hyperlink"
                cellule_hyperlien.alignment = Alignment(wrap_text=True)

            print("pourcentage pondérer")
            print(résulta[6])
            print("pourcentage image")
            print(résulta[5])
            print("--------------------------------------------")


        # Ajuster la largeur de chaque colonne de B à K pour que le texte s'affiche entièrement
        for colonne_index in range(2, 30):  # Colonnes B à K
            lettre_colonne = get_column_letter(colonne_index)
            largeur_texte = max(len("2050002398634") for cellule in ws[f'{lettre_colonne}'])
            colonne = ws.column_dimensions[lettre_colonne]
            colonne.width = largeur_texte + 2

        # Coordonnées de la cellule libre
        x = 14
        y = cellule_libre.row

        cellule_image_positif = ws.cell(row=y + 1, column=x)

        nom, Rb_sku, Image_path, fournisseur, fournisseur_sku = self.DataBase.getproductRB(RB_procuct)
        image_positif = Image("./image/Rb/" + fournisseur + "/" + Rb_sku + "/image_1.png")
        image_positif.width = 100
        image_positif.height = 100

        # Insérer l'image en spécifiant les coordonnées de la cellule libre
        ws.add_image(image_positif, f'{cellule_image_positif.column_letter}{cellule_image_positif.row}')

        wb.save(self.nom_fichier)

    def récupérerCorespondance(self, Rb_product):

        correspondances = self.DataBase.getCorespondanceDatabase(Rb_product)

        def comparer(correspondance):
            return (correspondance[6], correspondance[5])  # Utilisez les index des attributs ici

        index = len(correspondances) - 1
        while index >= 0:
            if correspondances[index][6] <= 20:
                del correspondances[index]
            index -= 1

        résultas_trier = sorted(correspondances, key=comparer, reverse=True)

        résultas_trier = résultas_trier[:10]

        return résultas_trier

    def trouver_premiere_cellule_vide_dans_rangée(self,ws, numero_rangee):
        for colonne_index in range(1, 100):
            cellule = ws.cell(row=numero_rangee, column=colonne_index)
            if cellule.value is None:
                return cellule

        # Si aucune cellule vide n'est trouvée, retourner None
        return None

    def get_couleur_pourcentage(self, pourcentage):
        if pourcentage < 20:
            # Rouge pâle
            return PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        elif 20 <= pourcentage < 30:
            # Orange
            return PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
        elif 30 <= pourcentage < 45:
            # Jaune pâle
            return PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        elif 45 <= pourcentage < 70:
            # vert pâle
            return PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        else:
            # Vert pétant
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')


if __name__ == "__main__":
    # Création de l'instance de la classe driverDataSet
    maprésentation = presentationResultas()
    maprésentation.présenterCorespondance("RB-Ada-09")
    maprésentation.présenterCorespondance("RB-Lyn-17")

