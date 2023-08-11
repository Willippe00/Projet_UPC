import openpyxl

class ExcelParser:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.current_sheet = None

    def select_sheet(self, sheet_name):
        self.current_sheet = self.workbook[sheet_name]

    def get_cell_value(self, row, column):
        return self.current_sheet.cell(row=row, column=column).value

    def get_max_rows(self):
        return self.current_sheet.max_row

    def get_max_columns(self):
        return self.current_sheet.max_column

    def get_range_values(self, start_row, start_column, end_row, end_column):
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for column in range(start_column, end_column + 1):
                row_data.append(self.get_cell_value(row, column))
            data.append(row_data)
        return data

    def close(self):
        self.workbook.close()

class explorateurTableur:

    def __init__(self):
        self.parser = ExcelParser("./sheet01.xlsx")
        self.parser.select_sheet("Sheet1")
        print("la feuille a été trouver")

    def parcourir(self, min, max):
        # Déclaration d'une liste vide
        listeCode = []

        for i in range(min , max +1):
            code = self.parser.get_cell_value(i,1)
            listeCode.append(code)
        return listeCode

    def parcourirAll(self):
        max = self.parser.get_max_rows()
        return self.parcourir(0,max)


